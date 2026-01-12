# -*- coding: utf-8 -*-
"""
Created on Thu Dec 11 23:51:41 2025

@author: watie
"""

#import rclpy # pour excel avec python
#from rclpy.node import Node
#from geometry_msgs.msg import Point
#from sensor_msgs.msg import JointState
import time
from openpyxl import load_workbook, Workbook
import xlrd #get data depuis Excel 
import pandas as pd #tableaux python 
import numpy as np 
from scipy.optimize import fsolve
import matplotlib.pyplot as plt


#############
path = 'C:/Users/watie/Documents/INSA/MIQ_5/S9/Info_indus/PJ_pantographe/ana7.xlsx'

# --- Longueurs du mécanisme ---
l0 = 138
l1 = 80
l2 = 146.3
l3 = 146.3
l4 = 80

epsilon = 10


#############

def gettablelength(path : str, sheetindex : int):

    try:
        # Read the Excel file
        df = pd.read_excel(path, sheet_name=sheetindex)

        # Get the length of the table
        table_length = len(df)

        return table_length
    except Exception as e:
        # Log the error
        print(f"Error: {e} -- gettablelength")
        return 0



def modeleinverse(x : float, y : float):
    
    # --- Valeur initiale ---
    q_init = [90, 90]  # important pour choisir la bonne solution

    #--- Solveur ---
    # sol = fsolve(
    #     forward_equations,
    #     q_init,
    #     args=(x, y, l0, l1, l2, l3, l4),
    # )
    
    sol, info, ier, msg = fsolve(
        forward_equations,
        q_init,
        args=(x, y, l0, l1, l2, l3, l4),
        full_output=True
    )
    
    
    if ier != 1:
        raise RuntimeError(f"fsolve n'a pas convergé pour x={x}, y={y} : {msg}")

    q1, q4 = sol
     
    return q1, q4 

    


def forward_MGD(q1, q4, l0, l1, l2, l3, l4):
    """
    Modèle géométrique direct (coude en haut).
    Retourne (x3, y3).
    """

    # --- Position de A2 ---
    x2 = l1 * np.cos(q1)
    y2 = l1 * np.sin(q1)

    # --- Position de A4 ---
    x4 = l0 + l4 * np.cos(q4)
    y4 = l4 * np.sin(q4)

    # --- Triangle A2-A3-A4 ---
    dx = x4 - x2
    dy = y4 - y2
    d = np.sqrt(dx**2 + dy**2)

    # Orientation A2->A4
    phi = np.arctan2(dy, dx)

    # Angle interne gamma (loi des cosinus)
    gamma = np.arccos((l2**2 + d**2 - l3**2) / (2 * l2 * d))

    # Orientation A2->A3 (coude en haut)
    theta23 = phi + gamma

    # --- Position de A3 ---
    x3 = x2 + l2 * np.cos(theta23)
    y3 = y2 + l2 * np.sin(theta23)

    return x3, y3

def forward_equations(q, x_target, y_target, l0, l1, l2, l3, l4):

    q1, q4 = q

    x3, y3 = forward_MGD(q1, q4, l0, l1, l2, l3, l4)

    return [x3 - x_target, y3 - y_target]

def plot_mechanism(q1, q4, l0, l1, l2, l3, l4):
    """Affiche le mécanisme pour les angles q1 et q4."""

    x3, y3 = forward_MGD(q1, q4, l0, l1, l2, l3, l4)

    A1 = np.array([0, 0])
    A5 = np.array([l0, 0])
    A2 = np.array([l1*np.cos(q1), l1*np.sin(q1)])
    A4 = np.array([l0 + l4*np.cos(q4), l4*np.sin(q4)])

    plt.figure()
    plt.plot([A1[0],A2[0]],[A1[1],A2[1]],'r-',linewidth=2)
    plt.plot([A2[0],x3],[A2[1],y3],'r-',linewidth=2)

    plt.plot([A5[0],A4[0]],[A5[1],A4[1]],'b-',linewidth=2)
    plt.plot([A4[0],x3],[A4[1],y3],'b-',linewidth=2)

    plt.plot(x3,y3,'ko',markersize=8)
    plt.text(x3,y3," A3")

    plt.grid(True)
    plt.axis('equal')
    plt.title("Mécanisme – configuration trouvée")
    plt.xlabel("x")
    plt.ylabel("y")
    plt.show()


# Ouvrir le fichier excel
wb = load_workbook(path, data_only=True)
#print(wb.sheetnames)

#On ouvre le tableau feuille par feuille 
sheet_x  = wb["Feuil3"]
sheet_y  = wb["Feuil4"]
sheet_q1 = wb["Feuil1"]
sheet_q4 = wb["Feuil2"]

#Cree un tableau de test vide 
lignes = gettablelength(path, 0)
colonnes = 5
tabtest = np.zeros((lignes-20, colonnes))

#On remplit le tableau avec les données du tableau excel
#TEMPS
#print(lignes)
for i in range(0, lignes-20):
     tabtest[i, 0] = sheet_q1.cell(row=i+22, column=1).value 
#q1
for i in range(0, lignes-20):
     tabtest[i, 1] = sheet_q1.cell(row=i+22, column=2).value  
#q4
for i in range(0, lignes-20):
     tabtest[i, 2] = sheet_q4.cell(row=i+22, column=2).value 
#x_eff
for i in range(0, lignes-20):
     tabtest[i, 3] = sheet_x.cell(row=i+22, column=2).value 
#y_eff
for i in range(0, lignes-20):
     tabtest[i, 4] = sheet_y.cell(row=i+22, column=2).value  


#Compare, pour des valeurs x, y donnees
#les valeurs q1, q4 correspondantes (théorie) en simulation
#avec les valeurs q1, q4 (experimentales) de notre modele inverse
#avec une marge d'erreur de EPSILON degrés

for i in range(0, lignes-20):
     x = tabtest[i, 3]
     y = tabtest[i, 4]
     q1_test, q4_test = modeleinverse(x, y)
     # print(x, y, q1_test, q4_test)
     # print(0,0, tabtest[i, 1],tabtest[i, 2])
     if(abs(q1_test - tabtest[i, 1]) > epsilon) or (abs(q4_test - tabtest[i, 2]) > epsilon):
         test = False
         print("test error ligne ",i, "\n La position q1, q4 des moteurs, calculés pour une position de l'effecteur en x: ", x, ", y : ", y, "n'est pas correcte" )
         break
     
if test:
    print("Tous les tests sont corrects")
        

        

            
             
         
